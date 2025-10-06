# Importando as bibliotecas 
import win32com.client
import pythoncom
import subprocess
import time
import datetime as dt
import keyring
import pandas as pd
from openpyxl import load_workbook
import logging

# Responsável por orquestrar a automação SAP
class mm:
    # Inicia a sessão com as configurações do SAP(usuário, ambiente e caminho do saplogon)
    def __init__(self, sap_user: str, sap_environment: str, sap_logon_path: str):
        """
        Args:
            sap_user (str): O nome de usuário SAP.
            sap_environment (str): O nome exato da conexão no SAP Logon.
            sap_logon_path (str): O caminho para o executável saplogon.
        """
        self.user = sap_user
        self.environment = sap_environment
        self.sap_path = sap_logon_path
        self.session = None

    # Tenta reutilizar uma conexão do SAP já aberta
    def _encontra_sessao_existente(self):
            """
            Verifica se já existe uma sessão SAP aberta para o ambiente especificado.

            Returns:
                session: O objeto de sessão se encontrado, senão None.
            """
            try:
                # Acessa o objeto de scripting do SAP GUI
                SapGuiAuto = win32com.client.GetObject("SAPGUI")
                if not SapGuiAuto:
                    return None

                application = SapGuiAuto.GetScriptingEngine
                if not application:
                    return None
                
                # Itera sobre as conexões abertas
                for i in range(application.Children.Count):
                    connection = application.Children(i)
                    # Verifica se o nome do ambiente corresponde
                    if self.environment in connection.Description:
                        session = connection.Children(0) # Pega a primeira sessão da conexão
                        logging.info(f"Sessão SAP encontrada para o ambiente '{self.environment}'. Reutilizando.")
                        return session
                return None
            except Exception:
                # Ocorre se o SAP GUI não estiver em execução ou não for acessível
                logging.info("Nenhuma instância do SAP GUI encontrada.")
                return None    
            
    # Abre o executável do SAP, caso não haja sessão disponível
    def _inicia_sap_gui(self):
        """Inicia o processo do SAP GUI."""
        try:
            subprocess.Popen(self.sap_path)
            time.sleep(3)  # Aguarda 3 segundos para o processo iniciar
            logging.info("SAP GUI iniciado com sucesso.")
        except Exception as e:
            logging.error(f"Falha ao iniciar o SAP GUI em '{self.sap_path}': {e}")
            raise   

    # Cria conexão com SAP, preenchendo usuário e senha

    # Tenta encontra sessão existente; caso contrário, efetua novo acesso
    def _conecta(self):
        """
        Ponto de entrada principal para obter uma sessão SAP.

        Verifica se uma sessão já existe. Se não, cria uma nova.
        """
        
        # Inicializa o ambiente COM (Component Object Model) que interage com o Excel e SAP GUI Scripting, por exemplo.
        pythoncom.CoInitialize()

        # Verifica se já existe uma sessão aberta e a retorna
        self.session = self._encontra_sessao_existente()
        if self.session:
            return self.session

        # Se não houver sessão, inicia o processo de login
        logging.info(f"Nenhuma sessão encontrada para '{self.environment}'. Iniciando novo processo de login.")
        self.session = self._novo_login()
        return self.session
    
    # Lê a planilha validada e ejusta os dados no formato esperado pelo SAP, principalmente datas e campos de texto
    def _relatorio(self, arquivo):    
        """
        Lê e prepara os dados da planilha já validada.
        
        Args:
            arquivo: o caminho da pasta com o nome da planilha validada.
            
        """
        # Padroniza e calcula datas auxiliares
        hoje = dt.date.today().strftime("%d/%m/%Y")
        dia = dt.date.today()
        tdelta = dt.timedelta(days=30)
        mes = dia + tdelta
        mes = mes.strftime("%d.%m.%Y")
        # Abre a planilha
        wb = load_workbook(arquivo)
        ws = wb.active
        # Lê dados da planilha
        data = pd.read_excel(arquivo, dtype=str)
        # Normaliza colunas: datas no formato dd.MM.yyyy, remove pontuação, troca ponto por vírgula,
        # converte o formato para string
        data['CNPJ_Fornecedor']  = data['CNPJ_Fornecedor'].astype(str)
        data['Data Emissao'] = pd.to_datetime(data['Data Emissao'].astype(str))
        data['Data Emissao'] = data['Data Emissao'].dt.strftime('%d.%m.%Y')
        data['Data In'] = pd.to_datetime(data['Data In'].astype(str))
        data['Data In'] = data['Data In'].dt.strftime('%d.%m.%Y')
        data['Data Out'] = pd.to_datetime(data['Data Out'].astype(str))
        data['Data Out'] = data['Data Out'].dt.strftime('%d.%m.%Y')
        data['Matricula']=data['Matricula'].str.replace('.','').replace('-','')
        data['Matricula']=data['Matricula'].str.replace('-','')
        data['Liquido a Pagar']  = data['Liquido a Pagar'].astype(str)
        data['Liquido a Pagar'] = data['Liquido a Pagar'].str.replace('.', ',')
        data['SST']  = data['SST'].astype(str)
        # Seleciona colunas essenciais para efetuar a automação 
        selecao = data[['CNPJ_Fornecedor', 'Data Emissao', 'RC', 'N° LINHA DA RC', 
                        'Matricula', 'Passageiro', 'Data In', 'Data Out', 'Requisicao de Viagem',
                        'Reserva de Recurso', 'Nota fiscal', 'Centro de Custo', 'Liquido a Pagar',
                        'PC', 'Fornecedor','DOMICILIO', 'FRS', 'SST']] 
        # Imprime as cinco últimas linhas da planilha para uma inspeção rápida
        print(selecao.tail())
        # Retorna uma lista
        lista = selecao.values.tolist()
        return lista
    
    # Cria requisições com base na lista de dados e as salva na planilha
    def _requisicao(self, lista, arquivo):
        """
        Gera requisição.
            
        Args:
            lista: é a principal fonte de dados para preenchimento dos campos no SAP. 
            arquivo: o caminho da pasta com o nome da planilha a ser atualizada, conforme execução do script.

        """

        # Verifica a sessão disponível, maximiza a janela e abre a ME51N (gera requisições)
        session = self.session
        if not session:
            logging.error("Sessão não disponível para _requisicao.")
            return
        session.findById("wnd[0]").maximize()
        session.starttransaction("ME51n")
        session.findById("wnd[0]").sendVKey(0)
        session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON").press()
        id = 1 #contador de item        
        codcusto = session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell")
        # Percorre a lista(cada j é uma linha da planilha)
        for i, j  in enumerate(lista):                            
        # Primeiro item da requisição
            if id == 1:                  
                # Preenche dados no SAP, se for K (Centro de Custo)
                if len(j[11]) == 7:             
            
                    codcusto.modifyCell(i,"BNFPO",id)
                    codcusto.modifyCell(i,"KNTTP","K")
                    codcusto.modifyCell(i,"EKGRP","F85")
                    codcusto.modifyCell(i,"TXZ01","HOSPEDAGEM NF {}".format(j[10]))
                    codcusto.modifyCell(i,"BEDNR",j[10])
                    codcusto.modifyCell(i,"WGBEZ","094300")
                    codcusto.modifyCell(i,"EPSTP","D")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "EPSTP"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "MEINS"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                    time.sleep(1)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-KTEXT1[1,0]").text = "HOSPEDAGEM NF {}".format(j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[2,0]").text = "1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "un"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[3,0]").text = j[12]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").caretPosition = 2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-KOSTL").text = j[11]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-KOSTL").caretPosition = 7
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").text = "8.8"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").caretPosition = 3
                    session.findById("wnd[0]").sendVKey(0)

                # Preenche dados do item, se for N (Ordem e Operação)
                elif j[11][0] == '1':

                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BNFPO", id)
                    codcusto.modifyCell(i,"KNTTP","N")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EKGRP","F85")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"TXZ01","HOSPEDAGEM NF {}".format(j[10]))
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BEDNR", j[10])
                    codcusto.modifyCell(i,"WGBEZ","094300")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EPSTP","D")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "EPSTP"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "MEINS"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                    time.sleep(1)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-KTEXT1[1,0]").text = "HOSPEDAGEM NF {}".format(j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[2,0]").text = "1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "UN"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[3,0]").text = j[12]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").caretPosition = 2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-NPLNR").text = j[11][0:10]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-VORNR").text = j[11][-4:]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-VORNR").setFocus()
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-VORNR").caretPosition = 2
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").text = "8.8"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").caretPosition = 3
                    session.findById("wnd[0]").sendVKey(0)
                
                # Preenche dados do item, se for P (Projeto)
                else:

                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BNFPO", id)
                    codcusto.modifyCell(i,"KNTTP","P")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EKGRP","F85")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"TXZ01", "HOSPEDAGEM NF {}".format(j[10]))
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BEDNR", j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"WGBEZ","094300")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EPSTP","D")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "EPSTP"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "MEINS"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                    time.sleep(1)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-KTEXT1[1,0]").text = "HOSPEDAGEM NF {}".format(j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[2,0]").text = "1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "UN"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[3,0]").text = j[12]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").setFocus
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").caretPosition = 2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-PS_POSID").text = j[11]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-PS_POSID").setFocus()
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-PS_POSID").caretPosition = 0
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").text = "8.8"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").caretPosition = 3
                    session.findById("wnd[0]").sendVKey(0)
                    
            # Percorre a lista, a partir do segundo item da requisição
            else:               

                codcust1 = session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell")

                # Preenche dados no SAP, se for K (Centro de Custo)  
                if len(j[11]) == 7: 

                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BNFPO",id)
                    codcust1.modifyCell(i,"KNTTP","K")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EKGRP","F85")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"TXZ01","HOSPEDAGEM NF {}".format(j[10]))
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BEDNR", j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EPSTP","D")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").setCurrentCell(i,"EPSTP")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "MEINS"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                    time.sleep(1)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-KTEXT1[1,0]").text = "HOSPEDAGEM NF {}".format(j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[2,0]").text = "1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "un"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[3,0]").text = j[12]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").caretPosition = 2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-KOSTL").text = j[11]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-KOSTL").caretPosition = 7
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").text = "8.8"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").caretPosition = 3
                    session.findById("wnd[0]").sendVKey(0)

                # Preenche dados do item, se for N (Ordem e Operação)
                elif j[11][0] == '1':              

                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BNFPO",id)
                    codcust1.modifyCell(i,"KNTTP","N")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EKGRP","F85")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"TXZ01","HOSPEDAGEM NF {}".format(j[10]))
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BEDNR", j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EPSTP","D")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").setCurrentCell(i,"EPSTP")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "MEINS"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                    time.sleep(1)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-KTEXT1[1,0]").text = "HOSPEDAGEM NF {}".format(j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[2,0]").text = "1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "UN"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[3,0]").text = j[12]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").caretPosition = 2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-NPLNR").text = j[11][0:10]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-VORNR").text = j[11][-4:]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-VORNR").setFocus()
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-VORNR").caretPosition = 4
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").text = "8.8"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").caretPosition = 3
                    session.findById("wnd[0]").sendVKey(0)
                    
                # Preenche dados do item, se for P (Projeto)
                else:

                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BNFPO",id)
                    codcust1.modifyCell(i,"KNTTP","P")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EKGRP","F85")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"TXZ01","HOSPEDAGEM NF {}".format(j[10]))
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"BEDNR",j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell(i,"EPSTP","D")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").setCurrentCell(i,"EPSTP")
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "MEINS"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                    time.sleep(1)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-KTEXT1[1,0]").text = "HOSPEDAGEM NF {}".format(j[10])
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[2,0]").text = "1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "UN"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[3,0]").text = j[12]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").setFocus
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").caretPosition = 2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-PS_POSID").text = j[11]
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-PS_POSID").setFocus()
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-PS_POSID").caretPosition = 24
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").text = "8.8"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT15/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1318/ssubCUSTOMER_DATA_ITEM:SAPLXM02:0111/tabsTABSTRIP_0111/tabpTAB3/ssubSUB03:SAPLXM02:1070/ctxtEBAN_CI-ZZTPCOD_TLC").caretPosition = 3
                    session.findById("wnd[0]").sendVKey(0)
                    
            # Incrementa o item para próxima linha
            id+=1

        # Após inserir os itens, salva a requisição e extrai  o número gerado na barra de status, preenchendo e 
        # gravando nas colunas AT (número da requisição), AU (número do item), AS (data da criação)
        # e AV (data da conclusão) da planilha
        session.findById("wnd[0]/tbar[0]/btn[11]").press()
        poCode = session.findById("wnd[0]/sbar").text
        poCode = poCode.split()
        poCode = poCode[6]
        poCode = int(poCode)
        print('RC nº {}'.format(poCode))
        wb = load_workbook(arquivo)
        ws = wb.active
        hoje = dt.date.today().strftime("%d/%m/%Y")

        for i in range(2, ws.max_row+1):
            ws['AT'+str(i)].value = poCode

        for i in range(2, ws.max_row+1):
            ws['AU'+str(i)].value = i-1
            ws['AS'+str(i)].value = hoje
            ws['AV'+str(i)].value = hoje


        wb.save(arquivo)
    print("Script finalizado")

    # Função auxiliar que trata o leiaute dinâmico (id_1 e id_2) da tela da transação ME21N,  
    # impedindo assim erro de execução do scrit de Criação de Pedido. 
    def encontrar_elemento(self, id_1, id_2):
        
        # Identifica a sessão disponível    
        session = self.session

        """
        Tenta encontrar um elemento na tela do SAP usando dois IDs possíveis.

        Args:
            session: O objeto de sessão ativa do SAP.
            id_1: A primeira variação do ID a ser tentada.
            id_2: A segunda variação do ID a ser tentada.

        Returns:
            O objeto de tela encontrado, ou None se nenhum ID funcionar.
        """
        if not session:
            logging.error("Sessão SAP não está ativa.")
            return None
        try:
            # Tenta encontrar com o primeiro ID
            elemento = session.findById(id_1)
            return elemento
        except:
            try:
                # Se o primeiro falhar, tenta com o segundo
                elemento = session.findById(id_2)
                return elemento
            except Exception as e:
                # Se ambos falharem, registra o erro e retorna None
                logging.error(f"Elemento não encontrado com os IDs '{id_1}' ou '{id_2}'.")
                return None
            
    # Cria pedidos com base na lista de dados e os grava na planilha
    def _pedido(self, lista, arquivo):
        """
        Gera pedido.
            
        Args:
            lista: é a principal fonte de dados para preenchimento dos campos no SAP. 
            arquivo: o caminho da pasta com o nome da planilha a ser atualizada, conforme execução do script.

        """
    # Verifica a sessão disponível, maximiza a janela e abre a ME51N (gera requisições)
        session = self.session
        if not session:
            logging.error("Sessão não disponível para _pedido.")
            return
        
        # Percorre a lista de dados das planilha, maximiza a janela e abre a ME51N (gera requisições)
        for i, j  in enumerate(lista):

                try:

                    session.findById("wnd[0]").maximize()
                    session.findById("wnd[0]/tbar[0]/okcd").text = "/NME21N"
                    session.findById("wnd[0]").sendVKey(0)
                    # IDs possíveis para o superfield
                    id_superfield_13 = "wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-SUPERFIELD"
                    id_superfield_16 = "wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-SUPERFIELD"

                    
                    # Encontra o elemento UMA VEZ do leiaute dinâmico
                    campo_superfield = self.encontrar_elemento(id_superfield_13, id_superfield_16)

                    # Realiza TODAS as ações na variável
                    if campo_superfield:
                        campo_superfield.caretPosition = 0
                        campo_superfield.setFocus()
                    else:
                        logging.warning(f"Pedido {i}: Campo Superfield não encontrado. Pulando item.")
                        continue # Pula para o próximo item do loop
                    
                    session.findById("wnd[0]").sendVKey(4)
                    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/btnG_SELFLD_TAB-MORE[6,56]").press()
                    session.findById("wnd[2]/usr/tabsTAB_STRIP/tabpNOSV").select()
                    session.findById("wnd[2]/usr/tabsTAB_STRIP/tabpNOSV/ssubSCREEN_HEADER:SAPLALDB:3030/tblSAPLALDBSINGLE_E/ctxtRSCSEL_255-SLOW_E[1,0]").text = "X"
                    session.findById("wnd[2]/usr/tabsTAB_STRIP/tabpNOSV/ssubSCREEN_HEADER:SAPLALDB:3030/tblSAPLALDBSINGLE_E/ctxtRSCSEL_255-SLOW_E[1,0]").caretPosition = 1
                    session.findById("wnd[2]/tbar[0]/btn[8]").press()
                    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[7,24]").text = j[0]
                    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[7,24]").setFocus()
                    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[7,24]").caretPosition = 14
                    session.findById("wnd[1]").sendVKey(0)
                    session.findById("wnd[1]").sendVKey(0)
                    
                    # IDs possíveis para o campo de data
                    id_data_13 = "wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-BEDAT"
                    id_data_16 = "wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-BEDAT"

                    # Encontra o elemento UMA VEZ
                    campo_data = self.encontrar_elemento(session, id_data_13, id_data_16)
                    print(campo_data)

                    # Realiza TODAS as ações na variável
                    if campo_data:
                        campo_data.text = j[1]
                        campo_data.setFocus()
                        campo_data.caretPosition = 2              
                    
                    # IDs possíveis para o botão
                    id_botao_13 = "wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON"
                    id_botao_16 = "wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON"

                    # Encontra o elemento UMA VEZ
                    botao_visao_geral = self.encontrar_elemento(session, id_botao_13, id_botao_16)

                    # Realiza a ação na variável
                    if botao_visao_geral:
                        botao_visao_geral.press()
                    
                    # Linhas de código extraídas do SAPScripting que navega em campos e telas do SAP
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT9/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1221/ctxtMEPO1222-EKGRP").text = "F85"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT9/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1221/ctxtMEPO1222-EKGRP").caretPosition = 3
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT1").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT1/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1226/ctxtMEPO1226-INCO1").text = "ZSE"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-BANFN[25,0]").text = j[2]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-BNFPO[26,0]").text = j[3]
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-BNFPO[26,0]").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-BNFPO[26,0]").caretPosition = 2
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT7/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1317/ctxtMEPO1317-MWSKZ").text = "D0"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT7/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1317/ctxtMEPO1317-MWSKZ").caretPosition = 2
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT6").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1313/txtMEPO1313-PLIFZ").text = "1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1313/txtMEPO1313-PLIFZ").caretPosition = 1
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT5").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT5/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1320/tblSAPLMEGUITC_1320/ctxtMEPO1320-EEIND[2,0]").text = mes
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT5/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1320/tblSAPLMEGUITC_1320/ctxtMEPO1320-EEIND[2,0]").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT5/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1320/tblSAPLMEGUITC_1320/ctxtMEPO1320-EEIND[2,0]").caretPosition = 10
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB1_0101/ssubSUB01:SAPLXM06:9101/ctxtEKKO_CI-ZZMODLICIT").text = "DP1"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB1_0101/ssubSUB01:SAPLXM06:9101/ctxtEKKO_CI-ZZMULTA").text = "0"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB1_0101/ssubSUB01:SAPLXM06:9101/ctxtEKKO_CI-ZZTPOBJ").text = "S"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB1_0101/ssubSUB01:SAPLXM06:9101/ctxtEKKO_CI-ZZBNAME").text = "SD0H"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB1_0101/ssubSUB01:SAPLXM06:9101/ctxtEKKO_CI-ZZMULTA").setFocus()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB1_0101/ssubSUB01:SAPLXM06:9101/ctxtEKKO_CI-ZZMULTA").caretPosition = 1
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB1_0101/ssubSUB01:SAPLXM06:9101/btnBT_GERFIS").press()
                    session.findById("wnd[1]/usr/btnBT_INSERT_FIS").press()
                    session.findById("wnd[2]/usr/ctxtEG_DADOS-CHAVE").text = "M359"
                    session.findById("wnd[2]/usr/ctxtEG_DADOS-CHAVE").caretPosition = 4
                    session.findById("wnd[2]/tbar[0]/btn[8]").press()
                    session.findById("wnd[1]/usr/btnBT_INSERT_FIS").press()
                    session.findById("wnd[2]/usr/ctxtEG_DADOS-CHAVE").text = "T3HV"
                    session.findById("wnd[2]/usr/ctxtEG_DADOS-CHAVE").caretPosition = 4
                    session.findById("wnd[2]/tbar[0]/btn[8]").press()
                    session.findById("wnd[1]/usr/btnBT_INSERT_FIS").press()
                    session.findById("wnd[2]/usr/ctxtEG_DADOS-CHAVE").text = "TFEX"
                    session.findById("wnd[2]/usr/ctxtEG_DADOS-CHAVE").caretPosition = 4
                    session.findById("wnd[2]/tbar[0]/btn[8]").press()
                    session.findById("wnd[1]/tbar[0]/btn[8]").press()
                    session.findById("wnd[1]/tbar[0]/btn[8]").press()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB4_0101").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB4_0101/ssubSUB04:SAPLXM06:9104/ctxtEKKO_CI-ZZTPCOD_TLC").text = "8.8"
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT11/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1227/ssubCUSTOMER_DATA_HEADER:SAPLXM06:0101/tabsTABSTRIP_0101/tabpTAB4_0101/ssubSUB04:SAPLXM06:9104/ctxtEKKO_CI-ZZTPCOD_TLC").caretPosition = 3
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT3").select()
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT3/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").text = f'{j[4]} - {j[5]} - {j[8]} - {j[6]} a {j[7]}'
                    
                    # Verifique que não há reserva de recursos, 
                    # salva o pedido e extrai  o número gerado na barra de status, preenchendo e 
                    # gravando nas colunas AY (número da pedido), AX (data da criação), AZ (data da conclusão)
                    # e BA (status) da planilha
                    if pd.isnull(j[9]):
    
                        session.findById("wnd[0]/tbar[0]/btn[11]").press()
                        pc = session.findById("wnd[0]/sbar").text
                        pc = pc.split()
                        st = pc[4]
                        pc = pc[8]
                        pc = int(pc)
                        print(pc)
                        hoje = dt.date.today().strftime("%d/%m/%Y")
                        ws['AY'+str(i+2)].value = pc
                        ws['AX'+str(i+2)].value = hoje
                        ws['AZ'+str(i+2)].value = hoje
                        ws['BA'+str(i+2)].value = st                
                        wb.save(arquivo)
                        
                        # Acessa a transação ME23N (Consulta Pedido), encontra documento fiscal correspondente na pasta,
                        # e o anexa no pedido
                        session.findById("wnd[0]/tbar[0]/okcd").text = "/NME23N"
                        session.findById("wnd[0]").sendVKey(0)
                        session.findById("wnd[0]/titl/shellcont/shell").pressContextButton("%GOS_TOOLBOX")
                        session.findById("wnd[0]/titl/shellcont/shell").selectContextMenuItem("%GOS_PCATTA_CREA")
                        session.findById("wnd[1]/usr/ctxtDY_PATH").text = caminho
                        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "NF {}.pdf".format(j[10])
                        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 9
                        session.findById("wnd[1]/tbar[0]/btn[0]").press()                  
                        

                    # Preenche o número da reserva de recursos, 
                    # salva o pedido e extrai  o número gerado na barra de status, preenchendo e 
                    # gravando nas colunas AY (número do pedido), AX (data da criação), AZ (data da conclusão)
                    # e BA (status) da planilha

                    else:
                        session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT3/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").setSelectionIndexes(9,9)
                        session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0010/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1").select()
                        session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/btnACCASS").press()
                        session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-KBLNR").text = j[9]
                        session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-KBLPOS").setFocus()
                        session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-KBLPOS").caretPosition = 0
                        session.findById("wnd[1]").sendVKey(4)
                        session.findById("wnd[2]/usr/tabsG_SELONETABSTRIP/tabpTAB003/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[2,24]").text = j[9]
                        session.findById("wnd[2]/usr/tabsG_SELONETABSTRIP/tabpTAB003/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/ctxtG_SELFLD_TAB-LOW[7,24]").text = "45510003"
                        session.findById("wnd[2]/usr/tabsG_SELONETABSTRIP/tabpTAB003/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/ctxtG_SELFLD_TAB-LOW[7,24]").setFocus()
                        session.findById("wnd[2]/usr/tabsG_SELONETABSTRIP/tabpTAB003/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/ctxtG_SELFLD_TAB-LOW[7,24]").caretPosition = 8
                        session.findById("wnd[2]").sendVKey(0)
                        session.findById("wnd[2]").sendVKey(0)
                        session.findById("wnd[1]").sendVKey(0)
                        session.findById("wnd[0]/tbar[0]/btn[11]").press()
                        session.findById("wnd[1]/usr/btnSPOP-VAROPTION1").press() 
                        pc = session.findById("wnd[0]/sbar").text
                        pc = pc.split()
                        st = pc[4]
                        pc = pc[8]
                        pc = int(pc)
                        print(pc)
                        ws['AY'+str(i+2)].value = pc
                        ws['AX'+str(i+2)].value = hoje
                        ws['AZ'+str(i+2)].value = hoje
                        ws['BA'+str(i+2)].value = st                
                        wb.save(arquivo)
                        
                        # Acessa a transação ME23N (Consulta Pedido), encontra documento fiscal correspondente na pasta,
                        # e o anexa no pedido
                        session.findById("wnd[0]/tbar[0]/okcd").text = "/NME23N"
                        session.findById("wnd[0]").sendVKey(0)
                        session.findById("wnd[0]/titl/shellcont/shell").pressContextButton("%GOS_TOOLBOX")
                        session.findById("wnd[0]/titl/shellcont/shell").selectContextMenuItem("%GOS_PCATTA_CREA")
                        session.findById("wnd[1]/usr/ctxtDY_PATH").text = caminho
                        session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "NF {}.pdf".format(j[10])
                        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 9
                        session.findById("wnd[1]/tbar[0]/btn[0]").press()

                except Exception as e:
                    print(f"Erro ao processar pedido {i}: {e}")            

        session.findById("wnd[0]/tbar[0]/btn[15]").press()
    print("Script finalizado")

    # Registra as folhas de serviço e as grava na planilha
    def _frs(self, lista, arquivo):
        
        """
        Gera folha de registro de serviço.
            
        Args:
            lista: é a principal fonte de dados para preenchimento dos campos no SAP. 
            arquivo: o caminho da pasta com o nome da planilha a ser atualizada, conforme execução do script.

        """
        # Identifica a sessão disponível
        session = self.session
        if not session:
            logging.error("Sessão não disponível para _frs.")
            return
        
        # Percorre a lista de dados das planilha, maximiza a janela e abre a Ml81N (gera as FRS)
        for i, j  in enumerate(lista):
            session.findById("wnd[0]").maximize()
            session.starttransaction("ML81N")
            session.findById("wnd[0]").sendVKey(0)
            
            # Linhas de código extraídas do SAPScripting que navega em campos e telas do SAP
            session.findById("wnd[1]/usr/ctxtRM11R-EBELN").text = j[13]
            session.findById("wnd[1]/usr/ctxtRM11R-EBELN").caretPosition = 10
            session.findById("wnd[1]").sendVKey(0)
            session.findById("wnd[0]/tbar[1]/btn[13]").press()
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGA").select()
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG").select()
            session.findById("wnd[0]/usr/txtESSR-TXZ01").text = "PGTO {}".format(j[14][:30])
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG/ssubSUB_HEADER:SAPLMLSR:0410/txtESSR-LBLNE").text = j[10]
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG/ssubSUB_HEADER:SAPLMLSR:0410/ctxtESSR-DLORT").text = j[15]
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG/ssubSUB_HEADER:SAPLMLSR:0410/ctxtESSR-LZVON").text = j[6]
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG/ssubSUB_HEADER:SAPLMLSR:0410/ctxtESSR-LZBIS").text = j[7]
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG/ssubSUB_HEADER:SAPLMLSR:0410/txtESSR-SBNAMAN").text = "SOLANO"
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGA/ssubSUB_ACCEPTANCE:SAPLMLSR:0420/ctxtESSR-BLDAT").text = j[1]
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGA/ssubSUB_ACCEPTANCE:SAPLMLSR:0420/txtESSR-XBLNR").text = j[10]
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGA/ssubSUB_ACCEPTANCE:SAPLMLSR:0420/txtESSR-BKTXT").text = "PGTO HOSPEDAGEM"
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG/ssubSUB_HEADER:SAPLMLSR:0410/txtESSR-SBNAMAN").setFocus()
            session.findById("wnd[0]/usr/tabsTAB_HEADER/tabpREGG/ssubSUB_HEADER:SAPLMLSR:0410/txtESSR-SBNAMAN").caretPosition = 6
            session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/btnSELEKTION").press()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/tbar[1]/btn[9]").press()
            session.findById("wnd[0]/tbar[0]/btn[11]").press()
            session.findById("wnd[1]/tbar[0]/btn[9]").press()        
            
            # Grava a FRS e extrai  o número gerado na barra de status, preenchendo e 
            # salvando nas colunas BB (número da FRS), BC (data da criação) e BD (data da conclusão)
            session.findById("wnd[1]/tbar[0]/btn[8]").press()
            frs = session.findById("wnd[0]/sbar").text
            frs = frs[31:42]
            frs = int(frs)
            print(frs)      
            ws['BB'+str(i+2)].value = frs
            ws['BC'+str(i+2)].value = hoje
            ws['BD'+str(i+2)].value = hoje
            wb.save(arquivo)
        session.findById("wnd[0]/tbar[0]/btn[15]").press()
    print('Script finalizado')

    # Registra os protocolos com a documentação  e os encaminha ao Setor Responsável para agendar o pagamento
    def _gd(self, lista, arquivo):
        
        """
        Gera protocolos de pagamento
            
        Args:
            lista: é a principal fonte de dados para preenchimento dos campos no SAP. 
            arquivo: o caminho da pasta com o nome da planilha a ser atualizada, conforme execução do script.

        """
            # Identifica a sessão disponível
            session = self.session
            if not session:
                logging.error("Sessão não disponível.")
                return
            # Percorre a lista de dados da planilha, maximiza a janela e abre a MlGD (gera protocolos)
            for i, j  in enumerate(lista):
                session.findById("wnd[0]").maximize()
                session.starttransaction("MLGD")
                session.findById("wnd[0]").sendVKey(0)
                
                # Linhas de código extraídas do SAPScripting que navegam em campos e telas do SAP
                session.findById("wnd[0]/usr/radRB_NF_SERVICO").setFocus()
                session.findById("wnd[0]/usr/radRB_NF_SERVICO").select()                        
                session.findById("wnd[0]/usr/txtV_SF_TOMA").text = "000111"
                session.findById("wnd[0]/usr/txtV_NFS").text = j[10]
                session.findById("wnd[0]/usr/ctxtW_PROTCAB-BLDAT").text = j[1]
                session.findById("wnd[0]/usr/ctxtW_PROTCAB-STCD1").text = j[0]
                session.findById("wnd[0]/usr/ctxtW_PROTCAB-TXJCD").text = j[15]
                session.findById("wnd[0]/usr/ctxtGV_FRS").text = j[16]
                session.findById("wnd[0]/usr/ctxtGV_FRS").setFocus()
                session.findById("wnd[0]/usr/ctxtGV_FRS").caretPosition = 10
                session.findById("wnd[0]/tbar[1]/btn[8]").press()
                session.findById("wnd[1]/usr/btnBT_SIM").press()
                session.findById("wnd[1]/usr/radRB_LOCAL").select()
                session.findById("wnd[1]/usr/radRB_LOCAL").setFocus()
                session.findById("wnd[1]/usr/btnBT_OK").press()
                session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
                session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = 0
                session.findById("wnd[1]").sendVKey(4)
                #Localiza a pasta onde fica o documento fiscal e o anexa ao protocolo
                session.findById("wnd[2]/usr/ctxtDY_PATH").text = pastaNF
                session.findById("wnd[2]/usr/ctxtDY_FILENAME").text = "NF {}.pdf".format(j[10])
                session.findById("wnd[2]/usr/ctxtDY_FILENAME").caretPosition = 13
                session.findById("wnd[2]/tbar[0]/btn[0]").press()
                # Grava o protocolo e extrai  o número gerado na barra de status, preenchendo e 
                # salvando nas colunas BF (número do protocolo), BC (data da criação) e BD (data da conclusão)
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
                gd = session.findById("wnd[0]/sbar").text
                GD = gd[10:20]
                GD = int(GD)
                print(GD)        
                ws['BF'+str(i+2)].value = GD
                ws['BG'+str(i+2)].value = hoje
                ws['BH'+str(i+2)].value = hoje
                wb.save(arquivo)
                
                # Acessa a transação MLGDC (Consulta protoco), encontra folha de registro de serviço correspondente na pastaFRS,
                # e o anexa ao protocolo
                session.starttransaction("MLGDC")
                session.findById("wnd[0]").sendVKey(0)
                session.findById("wnd[0]/usr/ctxtSO_BUKRS-LOW").text = "01"
                session.findById("wnd[0]/usr/ctxtSO_PROTC-LOW").text = GD
                session.findById("wnd[0]/usr/ctxtSO_PROTC-LOW").setFocus()
                session.findById("wnd[0]/usr/ctxtSO_PROTC-LOW").caretPosition = 10
                session.findById("wnd[0]/tbar[1]/btn[8]").press()
                session.findById("wnd[0]/usr/shell").selectedRows = "0"
                session.findById("wnd[0]/tbar[1]/btn[13]").press()
                session.findById("wnd[1]/usr/radRB_LOCAL").select()
                session.findById("wnd[1]/usr/radRB_LOCAL").setFocus()
                session.findById("wnd[1]/usr/btnBT_OK").press()
                session.findById("wnd[1]/usr/ctxtDY_PATH").text = pastaFRS
                session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "FRS {}.pdf".format(j[10])
                session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 11
                session.findById("wnd[1]/tbar[0]/btn[0]").press()            
            session.findById("wnd[0]/tbar[0]/btn[15]").press
    print("Script finalizado")